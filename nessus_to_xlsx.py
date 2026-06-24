#!/usr/bin/env python3
"""
nessus_to_xlsx.py — Convert a .nessus scan file into a formatted vulnerability report spreadsheet.

Usage:
    python3 nessus_to_xlsx.py <scan.nessus> [output.xlsx]

Requirements:
    pip install openpyxl pandas matplotlib

Output spreadsheet tabs:
    1. Executive Summary    — Key metrics, severity pie chart, top vulns bar chart
    2. Critical & High      — Prioritized remediation list (Critical + High only)
    3. All Findings         — Full dataset, filterable
    4. Hosts Summary        — Per-host vuln counts sorted by risk
    5. Remediation Priorities — Deduplicated vuln list grouped for patching
"""

import sys
import os
import xml.etree.ElementTree as ET
import pandas as pd
import tempfile

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════
# Color / style constants
# ═══════════════════════════════════════════════
DARK_NAVY   = '1B2A4A'
HEADER_BG   = '2C3E6B'
CRIT_BG     = 'C0392B'
HIGH_BG     = 'E67E22'
MED_BG      = 'F1C40F'
LOW_BG      = '3498DB'
WHITE       = 'FFFFFF'
LIGHT_GRAY  = 'F5F6FA'
BORDER_CLR  = 'D5D8DC'

SEV_MAP     = {'0': 'Informational', '1': 'Low', '2': 'Medium', '3': 'High', '4': 'Critical'}
SEV_ORDER   = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1}

SEV_FILLS = {
    'Critical': PatternFill('solid', fgColor=CRIT_BG),
    'High':     PatternFill('solid', fgColor=HIGH_BG),
    'Medium':   PatternFill('solid', fgColor=MED_BG),
    'Low':      PatternFill('solid', fgColor=LOW_BG),
}
SEV_FONTS = {
    'Critical': Font(name='Arial', bold=True, color=WHITE, size=10),
    'High':     Font(name='Arial', bold=True, color=WHITE, size=10),
    'Medium':   Font(name='Arial', bold=True, color=DARK_NAVY, size=10),
    'Low':      Font(name='Arial', bold=True, color=WHITE, size=10),
}

HEADER_FONT = Font(name='Arial', bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill('solid', fgColor=HEADER_BG)
DATA_FONT   = Font(name='Arial', size=10)
TITLE_FONT  = Font(name='Arial', bold=True, color=DARK_NAVY, size=14)
SUBTITLE_FONT = Font(name='Arial', bold=True, color=DARK_NAVY, size=12)
THIN_BORDER = Border(
    left=Side(style='thin', color=BORDER_CLR),
    right=Side(style='thin', color=BORDER_CLR),
    top=Side(style='thin', color=BORDER_CLR),
    bottom=Side(style='thin', color=BORDER_CLR),
)


# ═══════════════════════════════════════════════
# Helper functions
# ═══════════════════════════════════════════════
def style_header_row(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_data_cell(ws, row, col, wrap=False):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical='top', wrap_text=wrap)
    if row % 2 == 0:
        cell.fill = PatternFill('solid', fgColor=LIGHT_GRAY)


def apply_sev_style(ws, row, col, severity):
    if severity in SEV_FILLS:
        ws.cell(row=row, column=col).fill = SEV_FILLS[severity]
        ws.cell(row=row, column=col).font = SEV_FONTS[severity]


def shorten(name, n=40):
    return name if len(name) <= n else name[:n - 1].rsplit(' ', 1)[0] + '…'


# ═══════════════════════════════════════════════
# Parse .nessus XML into a DataFrame
# ═══════════════════════════════════════════════
def parse_nessus(filepath):
    tree = ET.parse(filepath)
    root = tree.getroot()

    total_hosts = len(root.findall('.//ReportHost'))
    report_name = 'Nessus Scan'
    for report in root.findall('.//Report'):
        report_name = report.get('name', report_name)

    rows = []
    for host in root.findall('.//ReportHost'):
        ip = host.get('name')
        props = {tag.get('name'): tag.text for tag in host.findall('.//HostProperties/tag')}
        hostname = props.get('host-fqdn', props.get('netbios-name', ''))
        host_os = props.get('operating-system', '')

        for item in host.findall('ReportItem'):
            sev = item.get('severity', '0')
            if sev == '0':
                continue
            cvss = item.findtext('cvss3_base_score', '') or item.findtext('cvss_base_score', '')
            rows.append({
                'IP Address':    ip,
                'Hostname':      hostname,
                'OS':            host_os,
                'Port':          item.get('port', ''),
                'Protocol':      item.get('protocol', ''),
                'Severity':      SEV_MAP.get(sev, sev),
                'Severity_Num':  int(sev),
                'CVSS Score':    cvss,
                'Plugin ID':     item.get('pluginID', ''),
                'Plugin Name':   item.get('pluginName', ''),
                'Plugin Family': item.get('pluginFamily', ''),
                'Synopsis':      item.findtext('synopsis', ''),
                'Description':   item.findtext('description', ''),
                'Solution':      item.findtext('solution', ''),
                'Risk Factor':   item.findtext('risk_factor', ''),
                'CVE':           ', '.join(c.text for c in item.findall('cve') if c.text),
                'See Also':      item.findtext('see_also', ''),
                'Plugin Output': (item.findtext('plugin_output', '') or '')[:500],
            })

    df = pd.DataFrame(rows)
    return df, total_hosts, report_name


# ═══════════════════════════════════════════════
# Generate the bar chart image with matplotlib
# ═══════════════════════════════════════════════
def make_bar_chart(df, output_path):
    top = (
        df.groupby(['Plugin ID', 'Plugin Name', 'Severity'])
        .agg(Hosts=('IP Address', 'nunique'))
        .reset_index()
    )
    top['sev_sort'] = top['Severity'].map(SEV_ORDER)
    top = top.sort_values(['sev_sort', 'Hosts'], ascending=[False, False]).head(15)

    labels = [f"{i+1}. {shorten(row['Plugin Name'])}" for i, (_, row) in enumerate(top.iterrows())]
    values = top['Hosts'].tolist()

    labels, values = labels[::-1], values[::-1]

    fig, ax = plt.subplots(figsize=(12, 8))
    bars = ax.barh(range(len(labels)), values, color='#2C3E6B', height=0.7)
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel('Affected Hosts', fontsize=11)
    ax.set_title('Top Vulnerabilities — Affected Host Count', fontsize=14, fontweight='bold', color='#1B2A4A')

    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                str(val), va='center', fontsize=10, fontweight='bold', color='#1B2A4A')

    ax.set_xlim(0, max(values) + 1)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()


# ═══════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════
def build_executive_summary(wb, df, total_hosts, report_name, chart_img_path):
    ws = wb.active
    ws.title = 'Executive Summary'
    ws.sheet_properties.tabColor = DARK_NAVY

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = f'Vulnerability Assessment — {report_name}'
    ws['A1'].font = Font(name='Arial', bold=True, color=DARK_NAVY, size=16)
    ws.merge_cells('A2:N2')
    ws['A2'] = 'Confidential'
    ws['A2'].font = Font(name='Arial', italic=True, color='7F8C8D', size=10)

    # Key Metrics (A-C, merged A:B for labels)
    ws['A4'] = 'Key Metrics'
    ws['A4'].font = SUBTITLE_FONT
    ws.merge_cells('A5:B5')
    ws.cell(row=5, column=1, value='Metric')
    ws.cell(row=5, column=3, value='Value')
    style_header_row(ws, 5, [1, 2, 3])

    stats = [
        ('Total Hosts Scanned',    total_hosts),
        ('Hosts with Findings',    df['IP Address'].nunique()),
        ('Total Findings',         len(df)),
        ('Critical',               len(df[df['Severity'] == 'Critical'])),
        ('High',                   len(df[df['Severity'] == 'High'])),
        ('Medium',                 len(df[df['Severity'] == 'Medium'])),
        ('Low',                    len(df[df['Severity'] == 'Low'])),
        ('Unique Vulnerabilities', df['Plugin ID'].nunique()),
    ]
    for i, (metric, val) in enumerate(stats, 6):
        ws.merge_cells(f'A{i}:B{i}')
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=1).font = DATA_FONT
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2).border = THIN_BORDER
        ws.cell(row=i, column=3, value=val)
        style_data_cell(ws, i, 3)
        if metric == 'Critical':
            ws.cell(row=i, column=3).fill = SEV_FILLS['Critical']
            ws.cell(row=i, column=3).font = Font(name='Arial', bold=True, color=WHITE)
        elif metric == 'High':
            ws.cell(row=i, column=3).fill = SEV_FILLS['High']
            ws.cell(row=i, column=3).font = Font(name='Arial', bold=True, color=WHITE)
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
            ws.cell(row=i, column=2).fill = PatternFill('solid', fgColor=LIGHT_GRAY)

    # Severity table (E-F) for pie chart
    ws['E4'] = 'Severity Distribution'
    ws['E4'].font = SUBTITLE_FONT
    ws.cell(row=5, column=5, value='Severity')
    ws.cell(row=5, column=6, value='Count')
    style_header_row(ws, 5, [5, 6])
    for i, sev in enumerate(['Critical', 'High', 'Medium', 'Low'], 6):
        count = len(df[df['Severity'] == sev])
        ws.cell(row=i, column=5, value=sev)
        ws.cell(row=i, column=6, value=count)
        apply_sev_style(ws, i, 5, sev)
        ws.cell(row=i, column=5).border = THIN_BORDER
        ws.cell(row=i, column=6).font = DATA_FONT
        ws.cell(row=i, column=6).border = THIN_BORDER

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10

    # Pie chart at H3
    pie = PieChart()
    pie.style = 10
    pie.add_data(Reference(ws, min_col=6, min_row=5, max_row=9), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=5, min_row=6, max_row=9))
    pie.width = 13
    pie.height = 10
    pie.title = None
    for i, color in enumerate(['C0392B', 'E67E22', 'F1C40F', '3498DB']):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    ws.add_chart(pie, 'H3')

    # Top vulnerabilities table (row 16+)
    TABLE_ROW = 16
    ws.merge_cells(f'A{TABLE_ROW}:F{TABLE_ROW}')
    ws[f'A{TABLE_ROW}'] = 'Top Critical & High Vulnerabilities by Affected Host Count'
    ws[f'A{TABLE_ROW}'].font = SUBTITLE_FONT

    HDR = TABLE_ROW + 1
    ws.merge_cells(f'A{HDR}:B{HDR}')
    ws.cell(row=HDR, column=1, value='Vulnerability')
    ws.cell(row=HDR, column=3, value='Severity')
    ws.cell(row=HDR, column=5, value='Hosts')
    ws.cell(row=HDR, column=6, value='Instances')
    style_header_row(ws, HDR, [1, 2, 3, 5, 6])

    top = (
        df.groupby(['Plugin ID', 'Plugin Name', 'Severity'])
        .agg(Hosts=('IP Address', 'nunique'), Instances=('IP Address', 'count'))
        .reset_index()
    )
    top['sev_sort'] = top['Severity'].map(SEV_ORDER)
    top = top.sort_values(['sev_sort', 'Hosts'], ascending=[False, False]).head(15)

    for r, (_, row) in enumerate(top.iterrows()):
        rr = HDR + 1 + r
        ws.merge_cells(f'A{rr}:B{rr}')
        ws.cell(row=rr, column=1, value=f"{r+1}. {row['Plugin Name']}")
        ws.cell(row=rr, column=1).font = DATA_FONT
        ws.cell(row=rr, column=1).border = THIN_BORDER
        ws.cell(row=rr, column=1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=rr, column=2).border = THIN_BORDER
        ws.cell(row=rr, column=3, value=row['Severity'])
        style_data_cell(ws, rr, 3)
        apply_sev_style(ws, rr, 3, row['Severity'])
        ws.cell(row=rr, column=5, value=row['Hosts'])
        style_data_cell(ws, rr, 5)
        ws.cell(row=rr, column=6, value=row['Instances'])
        style_data_cell(ws, rr, 6)
        if rr % 2 == 0:
            ws.cell(row=rr, column=1).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
            ws.cell(row=rr, column=2).fill = PatternFill('solid', fgColor=LIGHT_GRAY)

    last_data = HDR + len(top)

    # Bar chart image
    img = Image(chart_img_path)
    img.width = 900
    img.height = 600
    ws.add_image(img, f'A{last_data + 3}')


def build_critical_high(wb, df):
    ws = wb.create_sheet('Critical & High Findings')
    ws.sheet_properties.tabColor = 'C0392B'

    ws.merge_cells('A1:I1')
    ws['A1'] = 'Critical & High Findings — Prioritized for Remediation'
    ws['A1'].font = TITLE_FONT

    headers = ['Severity', 'CVSS', 'Plugin ID', 'Vulnerability', 'IP Address', 'Port', 'Synopsis', 'Solution', 'CVE']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, range(1, len(headers) + 1))

    ch = df[df['Severity'].isin(['Critical', 'High'])].sort_values(
        ['Severity_Num', 'CVSS Score'], ascending=[False, False]
    )
    for r, (_, row) in enumerate(ch.iterrows(), 3):
        vals = [row['Severity'], row['CVSS Score'], row['Plugin ID'], row['Plugin Name'],
                row['IP Address'], row['Port'], row['Synopsis'], row['Solution'], row['CVE']]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
            style_data_cell(ws, r, c, wrap=(c in [4, 7, 8, 9]))
        apply_sev_style(ws, r, 1, row['Severity'])

    widths = [12, 8, 12, 45, 16, 8, 40, 45, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A2:I{2 + len(ch)}"
    ws.freeze_panes = 'A3'


def build_all_findings(wb, df):
    ws = wb.create_sheet('All Findings')
    ws.sheet_properties.tabColor = '2C3E6B'

    headers = ['Severity', 'CVSS', 'Plugin ID', 'Vulnerability', 'Plugin Family',
               'IP Address', 'Hostname', 'Port', 'Protocol', 'Synopsis', 'Solution', 'CVE', 'Plugin Output']
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, range(1, len(headers) + 1))

    sdf = df.sort_values(['Severity_Num', 'CVSS Score'], ascending=[False, False])
    for r, (_, row) in enumerate(sdf.iterrows(), 2):
        vals = [row['Severity'], row['CVSS Score'], row['Plugin ID'], row['Plugin Name'],
                row['Plugin Family'], row['IP Address'], row['Hostname'], row['Port'],
                row['Protocol'], row['Synopsis'], row['Solution'], row['CVE'], row['Plugin Output']]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
            style_data_cell(ws, r, c, wrap=(c in [4, 10, 11, 12, 13]))
        apply_sev_style(ws, r, 1, row['Severity'])

    widths = [12, 8, 12, 40, 18, 16, 20, 8, 8, 40, 40, 25, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + len(df)}"
    ws.freeze_panes = 'A2'


def build_hosts_summary(wb, df):
    ws = wb.create_sheet('Hosts Summary')
    ws.sheet_properties.tabColor = 'E67E22'

    ws.merge_cells('A1:H1')
    ws['A1'] = 'Affected Hosts — Vulnerability Count by Severity'
    ws['A1'].font = TITLE_FONT

    headers = ['IP Address', 'Hostname', 'OS', 'Critical', 'High', 'Medium', 'Low', 'Total']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, range(1, len(headers) + 1))

    host_stats = []
    for ip in sorted(df['IP Address'].unique()):
        hdf = df[df['IP Address'] == ip]
        host_stats.append({
            'IP Address': ip,
            'Hostname':   hdf['Hostname'].iloc[0] if len(hdf) else '',
            'OS':         (hdf['OS'].iloc[0] or '')[:60] if len(hdf) else '',
            'Critical':   len(hdf[hdf['Severity'] == 'Critical']),
            'High':       len(hdf[hdf['Severity'] == 'High']),
            'Medium':     len(hdf[hdf['Severity'] == 'Medium']),
            'Low':        len(hdf[hdf['Severity'] == 'Low']),
            'Total':      len(hdf),
        })
    host_stats.sort(key=lambda x: (-x['Critical'], -x['High'], -x['Medium']))

    for r, hs in enumerate(host_stats, 3):
        for c, key in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=hs[key])
            style_data_cell(ws, r, c)
        if hs['Critical'] > 0:
            ws.cell(row=r, column=4).fill = SEV_FILLS['Critical']
            ws.cell(row=r, column=4).font = Font(name='Arial', bold=True, color=WHITE)
        if hs['High'] > 0:
            ws.cell(row=r, column=5).fill = SEV_FILLS['High']
            ws.cell(row=r, column=5).font = Font(name='Arial', bold=True, color=WHITE)

    widths = [16, 25, 40, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A2:H{2 + len(host_stats)}"
    ws.freeze_panes = 'A3'


def build_remediation(wb, df):
    ws = wb.create_sheet('Remediation Priorities')
    ws.sheet_properties.tabColor = '27AE60'

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Remediation Priorities — Grouped by Vulnerability'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:F2')
    ws['A2'] = 'Unique vulnerabilities sorted by severity and affected host count'
    ws['A2'].font = Font(name='Arial', italic=True, color='7F8C8D', size=10)

    headers = ['Severity', 'Plugin ID', 'Vulnerability', 'Affected Hosts', 'Solution', 'Affected IPs']
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, range(1, len(headers) + 1))

    remed = (
        df.groupby(['Plugin ID', 'Plugin Name', 'Severity', 'Solution'])
        .agg(Hosts=('IP Address', 'nunique'),
             IPs=('IP Address', lambda x: ', '.join(sorted(set(x)))))
        .reset_index()
    )
    remed['sev_sort'] = remed['Severity'].map(SEV_ORDER)
    remed = remed.sort_values(['sev_sort', 'Hosts'], ascending=[False, False])

    for r, (_, row) in enumerate(remed.iterrows(), 4):
        vals = [row['Severity'], row['Plugin ID'], row['Plugin Name'],
                row['Hosts'], row['Solution'], row['IPs']]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
            style_data_cell(ws, r, c, wrap=(c in [3, 5, 6]))
        apply_sev_style(ws, r, 1, row['Severity'])

    widths = [12, 12, 45, 14, 50, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(remed)}"
    ws.freeze_panes = 'A4'


# ═══════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    nessus_path = sys.argv[1]
    if not os.path.isfile(nessus_path):
        print(f"Error: file not found: {nessus_path}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base = os.path.splitext(os.path.basename(nessus_path))[0]
        output_path = f"{base}_Vulnerability_Report.xlsx"

    print(f"Parsing {nessus_path} ...")
    df, total_hosts, report_name = parse_nessus(nessus_path)
    print(f"  {total_hosts} hosts scanned, {len(df)} non-informational findings")
    print(f"  Critical: {len(df[df['Severity']=='Critical'])}  "
          f"High: {len(df[df['Severity']=='High'])}  "
          f"Medium: {len(df[df['Severity']=='Medium'])}  "
          f"Low: {len(df[df['Severity']=='Low'])}")

    if len(df) == 0:
        print("No findings to report.")
        sys.exit(0)

    print("Generating bar chart ...")
    chart_tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    chart_tmp.close()
    make_bar_chart(df, chart_tmp.name)

    print("Building spreadsheet ...")
    wb = Workbook()
    build_executive_summary(wb, df, total_hosts, report_name, chart_tmp.name)
    build_critical_high(wb, df)
    build_all_findings(wb, df)
    build_hosts_summary(wb, df)
    build_remediation(wb, df)

    wb.save(output_path)
    os.unlink(chart_tmp.name)
    print(f"Done → {output_path}")


if __name__ == '__main__':
    main()
